#Increment all prices by 10%


















'''
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet.cell(1,1)
print(cell.value)
for row in range (1, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    correct_price = cell.value * 2
    correct_price_cell = sheet.cell(row, 4)
    correct_price_cell.value = correct_price

values = Reference(sheet, min_row = 2, max_row = sheet.max_row,
                    min_col= 4,
                    max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')
'''